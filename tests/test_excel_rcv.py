from datetime import date
from openpyxl import load_workbook
from src.siat.excel_rcv import generar_excel_compras, COLUMNAS_RCV


def _tx(numero="162452780", importe="10.00"):
    return {"id": "TX-000001", "estado": "completado",
            "datos": {"nit": "1020703023", "numero_factura": numero,
                      "autorizacion": "123189FFD1971B", "fecha": "01/07/2026",
                      "importe": importe, "operadora": "ENTEL"}}


def test_cabeceras_coinciden_con_formulario():
    esperadas = ["Nº", "NIT Proveedor", "Código de Autorización", "Número Factura",
                 "Número DUI/DIM", "Fecha Factura", "Importe Total Compra",
                 "Importe ICE", "Importe IEHD", "Importe IPJ", "Tasas",
                 "Otro No Sujeto a Crédito Fiscal", "Importes Exentos",
                 "Importe Compras Gravadas a Tasa Cero", "Subtotal",
                 "Descuentos Bonificaciones y Rebajas Sujetas al IVA",
                 "Importe Gift Card", "Importe Base Crédito Fiscal", "Crédito Fiscal",
                 "Tipo Compra", "Código de Control", "Fecha de Registro"]
    assert COLUMNAS_RCV == esperadas


def test_fila_y_credito_fiscal(tmp_path):
    ruta = str(tmp_path / "c.xlsx")
    assert generar_excel_compras([_tx(importe="100.00")], ruta) == 1
    ws = load_workbook(ruta).active
    fila = {ws.cell(1, i + 1).value: ws.cell(2, i + 1).value for i in range(len(COLUMNAS_RCV))}
    assert str(fila["NIT Proveedor"]) == "1020703023"
    assert float(fila["Crédito Fiscal"]) == 13.00
    assert str(fila["Número DUI/DIM"]) == "0"
    assert str(fila["Tipo Compra"]) == "INTERNO/ACTIVIDAD"


def test_fecha_de_registro_es_hoy(tmp_path):
    ruta = str(tmp_path / "c.xlsx")
    generar_excel_compras([_tx()], ruta)
    ws = load_workbook(ruta).active
    idx = COLUMNAS_RCV.index("Fecha de Registro") + 1
    assert ws.cell(2, idx).value == date.today().strftime("%d/%m/%Y")
